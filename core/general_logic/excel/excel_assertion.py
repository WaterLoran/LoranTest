# coding: utf-8
import os
import jsonschema
import pandas as pd
import pytest_check
from openpyxl import load_workbook
from core.init import FILES_PATH

from core.logger.logger_interface import logger


class ExcelAssertion:

    def __init__(self):
        self.data_list = None

    def filter_by_condition(self, condition: list):
        column_name_line = self.data_list[0]
        row_count = len(self.data_list)

        filter_line = []

        for one_filter in condition:
            one_filter_line = []
            try:
                column_name = one_filter["column_name"]
                column_value = one_filter["column_value"]
            except TypeError:
                logger.error(
                    "你可能想用condition_simple 这种简化的表达式, 而不是condition, 请确认, 请搜索相关用例 test_excel_assert_001 和 使用condition_simple的demo")
                raise

            column_index = column_name_line.index(column_name)

            if not isinstance(column_value, list):
                column_value = [column_value]
            for i in range(row_count):
                cell = self.data_list[i][column_index]
                if cell in column_value:  # 表示这个值在目标过滤值中
                    one_filter_line.append(i)
            filter_line.append(one_filter_line)

        # 然后对 filter_line 中的所有列表, 进行取交集
        if len(filter_line) == 0:
            return []
        else:
            final_set = filter_line[0]
            for line in filter_line:
                set1 = set(final_set)
                set2 = set(line)
                final_set = set1 & set2
            return final_set

    def read_excel_of_xls_format(self, file_path="", sheet_name="Sheet1"):
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        data_list = []
        header = list(df.columns)
        data_list.append(header)

        # 遍历 DataFrame 中的每一行，并将每一行的数据存储在 data_list 中
        for index, row in df.iterrows():
            row_data = list(row)
            data_list.append(row_data)

        self.data_list = data_list

    def read_excel_of_xlsx_format(self, file_path="", sheet_name="Sheet1"):
        logger.debug("读取的文件的路径为" + file_path)
        workbook = load_workbook(file_path)
        worksheet = workbook[sheet_name]
        max_row = worksheet.max_row
        max_column = worksheet.max_column

        data_list = []
        for i in range(1, max_row + 1):
            line_data = []
            for j in range(1, max_column + 1):
                cell_data = worksheet.cell(i, j).value
                line_data.append(cell_data)
            data_list.append(line_data)
        self.data_list = data_list

    def change_condition_simple_to_condition(self, condition_simple):
        condition = []
        for line in condition_simple:
            condition.append(
                {
                    "column_name": line[0],
                    "column_value": line[1],
                }
            )
        return condition

    def judge_type_of_the_condition(self, check_data):
        # 定义一个 condition为字典列表的类型
        dict_list_schema = {
            "type": "object",
            "properties": {
                "file": {"type": "string"},
                "sheet": {"type": "string"},
                "assert_type": {"type": "string"},
                "condition": {
                    "type": "array",
                    "items": {
                        "type": "object"  # 表示里面的元素都是对象类型的, 即dict
                    }
                },
                "target": {  # 表示target字段是list类型, 或者整形
                    "anyOf": [
                        {"type": "integer"},
                        {"type": "array"}
                    ]
                },

            },
            "required": ["file", "condition", "assert_type", "target"]
        }

        # 定义一个 condition为列表列表的类型, 即condition是一个列表, 里面的元素也是列表
        list_list_schema = {
            "type": "object",
            "properties": {
                "file": {"type": "string"},
                "sheet": {"type": "string"},
                "assert_type": {"type": "string"},
                "condition": {
                    "type": "array",
                    "items": {
                        "type": "array"  # 表示里面的元素都是对象类型的, 即dict
                    }
                },
                "target": {  # 表示target字段是list类型, 或者整形
                    "anyOf": [
                        {"type": "integer"},
                        {"type": "array"}
                    ]
                },

            },
            "required": ["file", "condition", "assert_type", "target"]
        }

        try:
            jsonschema.validate(check_data, dict_list_schema)
            logger.debug("check_excel断言操作中, check中condition的格式为dict_list类型")
            condition_type = "dict_list"  # 表示condition中的所有元素的类型都是字典
            return condition_type
        except Exception as err:
            p_err1 = err
            pass

        try:
            jsonschema.validate(check_data, list_list_schema)
            logger.debug("check_excel断言操作中, check中condition的格式为list_list类型")
            condition_type = "list_list"  # 表示condition中的所有元素的类型都是字典
            return condition_type
        except Exception as err:
            p_err2 = err
            pass

        logger.error("尝试判断为dict_list类型的时候, 报错信息为" + str(p_err1))
        logger.error("尝试判断为list_list类型的时候, 报错信息为" + str(p_err2))
        raise

    def do_assert(self, check: dict):

        # 对check的入参做格式判断
        condition_type = self.judge_type_of_the_condition(check)

        file = check["file"]
        if "sheet" in check.keys():
            sheet = check["sheet"]
        else:
            sheet = "Sheet1"

        # ======================================================== 处理condition
        if condition_type == "dict_list":
            condition = check["condition"]
        elif condition_type == "list_list":
            condition_list_list = check["condition"]
            condition = self.change_condition_simple_to_condition(condition_list_list)
        else:
            raise
        # ======================================================== 处理condition

        assert_type = check["assert_type"]
        target = check["target"]

        # 根据file到执行目录去读取出信息, 并存入字典
        file_path = os.path.join(FILES_PATH, file)

        if file_path.endswith(".xls"):
            self.read_excel_of_xls_format(file_path=file_path, sheet_name=sheet)
        else:
            self.read_excel_of_xlsx_format(file_path=file_path, sheet_name=sheet)

        logger.debug("")
        if isinstance(condition, int):
            logger.error("condition为过滤条件, 不能输入整型, 只能在target中输入整形")
            raise
        condition_filter = self.filter_by_condition(condition)
        logger.debug(f"根据业务脚本层调用check_excel函数并传入的参数中的condition过滤所得的行 为 {condition_filter}")

        target_filter = {999999999999}  # 构造一个不会轻易有交集的集合

        if isinstance(target, int):  # 如果是int型号, 表示, 要做数量上的一个断言
            logger.debug("传入的target为整型数字, 即期望 通过condition过滤出来的行数量 等于target")
            target_count = target
        else:
            target_count = None
            target_filter = self.filter_by_condition(target)
            logger.debug(
                f"根据业务脚本层调用check_excel函数并传入的参数中的target过滤所得的行 为 {target_filter}")

        if target_count is not None:  # 即传入的是数字:
            if assert_type not in ["equal", "=="]:
                logger.error(
                    "当target为数字的时候, 即表示期望通过condition过滤出来的数量为{target}个, 并且断言方式目前仅涉及为支持 相等断言")
                raise
            else:
                logger.debug(
                    f"target为数字, 即期望 根据condition过滤所得的行数量为 {target}, 当前condition过滤所得为{condition_filter}")
                if pytest_check.equal(len(condition_filter), target_count):
                    logger.debug("根据condition过滤所得的行的数量 等于 target, 断言成功")
                else:
                    logger.debug("根据condition过滤所得的行的数量 不等于 target, 断言失败")
        else:
            if assert_type not in ["=>", "<=", "==", "equal", "=="]:
                logger.error("目前设计上仅仅支持 集合的相等, 包含, 和被包含 断言, A=>B 表示, A包含B")
                raise
            elif assert_type == "=>":
                logger.debug("期望 target所过滤的行集合 是 condition所过滤的行集合 的子集")
                if target_filter <= condition_filter:
                    logger.debug("集合关系断言成功")
                else:
                    logger.debug("excel_check中的集合关系断言失败")
                    pytest_check.is_false()
            elif assert_type == "<=":
                logger.debug("期望 condition所过滤的行集合 是 target所过滤的行集合 的子集")
                if condition_filter <= target_filter:
                    logger.debug("集合关系断言成功")
                else:
                    logger.debug("excel_check中的集合关系断言失败")
                    pytest_check.is_false()
                pass
            elif assert_type in ["equal", "=="]:
                logger.debug("期望 target所过滤的行集合 等于 condition所过滤的行集合")
                if pytest_check.equal(condition_filter, target_filter):
                    logger.debug("集合关系断言成功")
                else:
                    logger.debug("集合关系断言失败")
